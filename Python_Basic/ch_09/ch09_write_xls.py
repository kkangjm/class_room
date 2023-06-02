
# openpyxl 모듈 불러오기
from openpyxl import Workbook

# 엑셀 워크북 인스턴스 생성
write_wb = Workbook()

# 현재 활성화된 시트 선택
ws = write_wb.active

# 선택된 시트명 변경
ws.title = 'My_sheet_1'

# My_sheet_1 시트 선택 후, write_ws 객체 지정
write_ws = write_wb['My_sheet_1']

# 셀 값 변경
write_ws['A1'] = 'TEST'

# 행 단위로 추가
write_ws.append(['A','B','C'])
write_ws.append([1,2,3])

# 셀 단위로 추가
write_ws.cell(5, 5, '5행5열')

# My_sheet_2 시트 생성
write_ws2 = write_wb.create_sheet('My_sheet_2')

# 셀 단위로 추가
write_ws2.cell(2, 2, 'My_sheet2_data')

# 엑셀 워크북 저장
write_wb.save("xls_sample.xlsx")

