from matplotlib import pyplot as plt
from openpyxl import load_workbook


# K data 읽기
load_wb = load_workbook('K.xlsx', data_only=True)
load_ws = load_wb.active
range_data = load_ws['A']
time_k = []
data_k = []
for i in range(2,len(range_data)+1):
    time_k.append(load_ws.cell(i, 1).value)
    data_k.append(load_ws.cell(i, 2).value)

# S data 읽기
load_wb = load_workbook('S.xlsx', data_only=True)
load_ws = load_wb.active
range_data = load_ws['A']
time_s = []
data_s = []
for i in range(2,len(range_data)+1):
    time_s.append(load_ws.cell(i, 1).value)
    data_s.append(load_ws.cell(i, 2).value)


# K 그래프
fig, ax1 = plt.subplots(figsize=(8,4))
ax1.set_xlabel('time')
ax1.set_ylabel('K - Axis', color='blue')
line_k = ax1.plot(time_k, data_k, label='K Data', color='blue')

# Line형 그래프
ax2 = ax1.twinx()   # X축 공유
ax2.set_ylabel('S - Axis', color='red')
line_s = ax2.plot(time_s, data_s, label='S Data', color='red')

# 그리드 표시
plt.grid(True)

# 범례 표시
lines = line_k + line_s
labels = [l.get_label() for l in lines]
ax1.legend(lines, labels, loc='upper left')

# 그래그 그리기
plt.tight_layout()  # 그래프 레이아웃 채우기
plt.show()
