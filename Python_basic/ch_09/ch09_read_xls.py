
# openpyxl 모듈 불러오기
from openpyxl import load_workbook

# 엑셀 워크북 읽기 인스턴스 생성
load_wb = load_workbook("xls_sample.xlsx", data_only=True)

# 시트 이름으로 불러오기
load_ws = load_wb['My_sheet_1']

# 셀 주소로 값 출력
print(load_ws['A1'].value)

# 셀 좌표로 값 출력
print(load_ws.cell(5, 5).value)

# 지정한 셀의 값 출력
range_data = load_ws['A2': 'C3']
for row_data in range_data:
    for cell_data in row_data:
        print(cell_data.value)


