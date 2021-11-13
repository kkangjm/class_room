
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
import statistics

def cor_factor(x, y):
    avr_x = statistics.mean(x)
    avr_y = statistics.mean(y)
    sum = 0
    for i in range(len(x)):
        sum = sum + (x[i]-avr_x)*(y[i]-avr_y)
    covar = sum / len(x)
    cor_fac = covar / (statistics.pstdev(x)*statistics.pstdev(y))
    return cor_fac


# event loop 창 숨기기
Tk().withdraw()

# file open dialog 호출
file_X = askopenfilename(title="Select x data file", filetypes=(('xlsx files', '*.xlsx'),))
file_Y = askopenfilename(title="Select y data file", filetypes=(('xlsx files', '*.xlsx'),))

# X data 읽기
load_wb = load_workbook(file_X, data_only=True)
load_ws = load_wb.active
range_data = load_ws['A']
data_x = []
for i in range(1,len(range_data)+1):
    data_x.append(load_ws.cell(i,1).value)

# Y data 읽기
load_wb = load_workbook(file_Y, data_only=True)
load_ws = load_wb.active
range_data = load_ws['A']
data_y = []
for i in range(1,len(range_data)+1):
    data_y.append(load_ws.cell(i,1).value)

print('Correlation_factor = ', cor_factor(data_x, data_y))

